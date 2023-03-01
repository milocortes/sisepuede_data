from openpyxl import load_workbook
import pandas as pd
import string
import sys 
import numpy as np
import subprocess

## Recibimos argumentos del usuario

id_core = int(sys.argv[1])
n_cores = int(sys.argv[2])

## Nombre del archivo
archivo = "fuel-subsidies-template-2022.xlsx"

#### Cargamos el archivo como sólo lectura para obtener los datos de la pestaña lists
### Obten lista de Países
wb = load_workbook(archivo, data_only=True) 

sheet_name = "lists"

sheet_lists = wb[sheet_name]

# Países
raw_data = [(sheet_lists[f'A{i}'].value, sheet_lists[f'B{i}'].value) for i in range(2, 194)]
df_countries = pd.DataFrame(raw_data, columns = ["Countries", "Region"])

# Años
anios = pd.Series([int(sheet_lists[f"F{i}"].value) for i in range(2,13)])

# Unidades 
unidades = pd.Series([sheet_lists[f"H{i}"].value for i in range(2,5)])

# Cargamos datos de iso_code3

# Load ISO3 code
iso3_m49_correspondence = pd.read_html("https://unstats.un.org/unsd/methodology/m49/")[0]

iso3_m49_correspondence.rename(columns = {"Country or Area" : "Nation", "ISO-alpha3 code" : "iso_code3"}, inplace = True)
iso3_m49_correspondence = iso3_m49_correspondence[["Nation", "iso_code3"]]
iso3_m49_correspondence = iso3_m49_correspondence.rename(columns = {"Nation":"Countries"})

# Agregamos el iso a los datos de los paises
df_countries = df_countries.merge(right=iso3_m49_correspondence, how = 'left', on = 'Countries')

#### Cargamos el archivo manteniendo las fórmulas para poder hacer 
#### cambios en los inputs de la pestaña "Single_Country_Fuel" 
#### Haremos un loop para recorrer por país y por año

wb = load_workbook(archivo)

sheet_name = "Single_Country_Fuel" 

sheet_single_country_fuel = wb[sheet_name]


"""
### Cambia dato en los inputs:
    * Select country (D4)
    * Select year (D5)
"""

def valores_chunk(n, n_cores, id_core):
    chunk = n//n_cores

    rango_inicio = chunk*(id_core-1)

    if not id_core == n_cores:
        rango_final = rango_inicio + chunk
    else:
        rango_final = n 

    return range(rango_inicio, rango_final)


n = df_countries.shape[0]

my_chunk = valores_chunk(n, n_cores, id_core)

file_object = open(f"{id_core}_files_fuel_subsidies.txt", 'w')

for idx in my_chunk:
    
    country = df_countries.loc[idx,"Countries"]
    iso_code3_tag = df_countries.loc[idx,"iso_code3"]
    
    for anio in anios:
        print(f"Procesando pais {country} #{idx}, anio {anio}")
        sheet_single_country_fuel["D4"].value = country
        sheet_single_country_fuel["D5"].value = anio

        ### Guardamos el archivo con lo cambios realizados en los inputs

        if isinstance(iso_code3_tag, str):
            file_name = f'xlsx_files/{iso_code3_tag}-{anio}.xlsx'
            file_object.write(f'{iso_code3_tag}_{anio}.xlsx\n')
            csv_file_name = f"{iso_code3_tag}-{anio}.csv"

        else:
            country = country.replace(" ","_")
            file_name = f'xlsx_files/{country}-{anio}.xlsx'
            file_object.write(f'{country}_{anio}.xlsx\n')
            csv_file_name = f"{country}-{anio}.csv"

        wb.save(file_name)

        cmd_convert_to_csv = f"libreoffice --headless --convert-to csv {file_name}"
        subprocess.run(cmd_convert_to_csv, shell=True)

        cmd_rm_xlsx = f"mv {csv_file_name} csv_files/ && rm {file_name}"
        subprocess.run(cmd_rm_xlsx, shell=True)


file_object.close()
wb.close()